"""Extract benchmark-to-benchmark lineage critiques + external views via qwen-max.

Inputs:
  - 科技伦理toolkit.xlsx (Benchmark诊断 sheet) – existing per-rubric comments
  - scripts/lineage_map*.json – curated short_name/year/critiques_received/self_acknowledged_limits
  - A_Comprehensive_Survey_of_AI_Ethics_Benchmarks (6).pdf – external view source
  - 诊断rubrics说明 sheet – Survey来源 column references (Hancox-Li, Raji, Eriksson, Corrêa, Gallegos…)

Outputs:
  - scripts/lineage_views.raw.jsonl – raw Qwen responses
  - scripts/lineage_views.json – {benchmark_short_name: {...}}
  - 科技伦理toolkit_lineage与外部观点.xlsx – one row per benchmark

Design choice (per user spec): 评论尽量保持原文。Qwen 是组织者，不是改写器：
  - 输入 prompt 把 critiques_received 原句、self_acknowledged_limits 原句、Survey 段落原文都灌进去
  - Prompt 强约束：外部观点必须以引号 + 来源 引出，禁止 paraphrase
  - 已有诊断表的 rubric 评论原样保留，本表只做"lineage 与外部观点"补充
"""
from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

import urllib.request
import urllib.error
import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parent
TOOLKIT = ROOT.parent / "workbooks" / "科技伦理toolkit.xlsx"
SURVEY_PDF = ROOT.parent / "A_Comprehensive_Survey_of_AI_Ethics_Benchmarks (6).pdf"
OUT_XLSX = ROOT.parent / "workbooks" / "科技伦理toolkit_lineage与外部观点.xlsx"
OUT_JSON = ROOT / "lineage_views.json"
OUT_RAW = ROOT / "lineage_views.raw.jsonl"

KEY_FILE = Path.home() / ".dashscope_key"
api_key = KEY_FILE.read_text().strip()
DASHSCOPE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
MODEL = "qwen-max"

# --- 1. Load lineage maps -----------------------------------------------------
LINEAGE_FILES = {
    "fairness": ROOT / "lineage_map.json",
    "humancentered": ROOT / "lineage_map_humancentered.json",
    "privacy": ROOT / "lineage_map_privacy.json",
    "safety": ROOT / "lineage_map_safety.json",
    "trustworthy": ROOT / "lineage_map_trustworthy.json",
}

def load_lineage() -> dict:
    out = {}
    for src, path in LINEAGE_FILES.items():
        data = json.loads(path.read_text(encoding="utf-8"))
        for k, v in data.items():
            if k == "_doc" or not isinstance(v, dict) or v.get("_alias_of"):
                continue
            sname = v.get("short_name", k)
            out[sname] = {**v, "_pillar": src, "_fid": k}
    return out

# --- 2. Load existing diagnosis rows -----------------------------------------
def load_diagnosis() -> list[dict]:
    wb = openpyxl.load_workbook(TOOLKIT, data_only=True)
    ws = wb["Benchmark诊断"]
    rubric_names = [ws.cell(row=2, column=c).value for c in range(2, 16)]
    rows = []
    for r in range(3, ws.max_row + 1):
        cite = ws.cell(row=r, column=16).value
        if not cite:
            continue
        tag = ws.cell(row=r, column=1).value
        rubrics = {}
        for i, name in enumerate(rubric_names):
            v = ws.cell(row=r, column=2 + i).value
            if v:
                rubrics[name] = str(v)
        ds_note = ws.cell(row=r, column=17).value
        rows.append({
            "row": r,
            "tag": tag,
            "cite": str(cite).strip(),
            "rubrics": rubrics,
            "dataset_note": str(ds_note) if ds_note else "",
        })
    return rows

# --- 3. Survey PDF passage index ---------------------------------------------
def load_survey_text() -> str:
    cache = Path("/tmp/survey_full.txt")
    if cache.exists():
        return cache.read_text()
    with pdfplumber.open(SURVEY_PDF) as pdf:
        pages = []
        for p in pdf.pages:
            t = p.extract_text() or ""
            t = re.sub(r"^\s*\d+\s*$", "", t, flags=re.MULTILINE)
            t = re.sub(r"\n\s*\d{1,4}\s*\n", "\n", t)
            pages.append(t)
        full = "\n\n[PAGE BREAK]\n\n".join(pages)
    cache.write_text(full)
    return full

def find_survey_passages(text: str, needle: str, window: int = 400) -> list[str]:
    """Return short context windows around each occurrence of `needle`."""
    out, seen = [], set()
    for m in re.finditer(re.escape(needle), text, flags=re.IGNORECASE):
        s, e = max(0, m.start() - window), min(len(text), m.end() + window)
        # snap to sentence boundary roughly
        snippet = text[s:e].replace("\n", " ")
        snippet = re.sub(r"\s+", " ", snippet).strip()
        key = snippet[:80]
        if key in seen:
            continue
        seen.add(key)
        out.append(snippet)
    return out[:3]  # cap at 3 passages per benchmark

# --- 4. Cite-name → lineage / diagnosis row matcher --------------------------
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())

def build_match_index(rows: list[dict], lineage: dict) -> list[dict]:
    """For each diagnosis row with a benchmark Cite, attach lineage entry if matched."""
    norm_lineage = {normalize(k): k for k in lineage.keys()}
    bench_rows = []
    for row in rows:
        cite = row["cite"]
        # Rows 3-19 are the Survey/methodology block – not benchmarks, skip
        if row["row"] <= 19:
            bench_rows.append({**row, "is_survey": True, "lineage": None, "short_name": cite[:60]})
            continue
        # extract candidate short name – take leading part before "("
        head = re.split(r"\s*[\(（]", cite)[0].strip()
        # strip trailing tokens like " / "
        head = re.split(r"\s+/\s+", head)[0].strip()
        # Try direct + normalized match
        short = None
        if head in lineage:
            short = head
        else:
            nh = normalize(head)
            for nk, k in norm_lineage.items():
                if nk == nh or (len(nh) > 3 and (nh in nk or nk in nh)):
                    short = k
                    break
        bench_rows.append({**row, "is_survey": False, "lineage": lineage.get(short) if short else None,
                           "short_name": short or head})
    return bench_rows

# --- 5. Qwen prompt -----------------------------------------------------------
SYSTEM_PROMPT = """你是 AI 伦理 benchmark 的文献综述助手。你的任务是为单个 benchmark 整理"lineage 引用 + 外部文章观点"补充评论。

**最高原则：尽量保留原文，不要 paraphrase。**

输入会提供：
1. benchmark 名、所属 pillar、cite 字段
2. 已有 14-rubric 中文评论（仅作上下文，不要重写）
3. lineage_map 中的 critiques_received（后来论文对它的批评，原文）
4. lineage_map 中的 self_acknowledged_limits（作者自承局限，原文）
5. A_Comprehensive_Survey_of_AI_Ethics_Benchmarks 中提到本 benchmark 的段落（原文）

输出严格 JSON：
{
  "短名": "...",                     // benchmark 短名
  "lineage_critiques": [             // 后来 benchmark 对它的批评，每条都必须是 critiques_received 中的原句，可加引号
     "「原句」— 出处简称"
  ],
  "self_limits": "「原句」",         // self_acknowledged_limits 原句（带引号）；若无填 ""
  "external_views": [                // Survey/Hancox-Li/Raji/Eriksson 的原文片段，每条原句+来源
     "「Survey 第 X 段：原句」"
  ],
  "synthesis": "..."                 // ≤120 字中文小结：把上面三块串起来，可用'综述指出'/'后续工作批评'等连接词；其余必须是引用片段。
}

规则：
- 若 critiques_received/self_acknowledged_limits/Survey 段落为空，则对应字段为空 [] 或 ""。
- external_views 必须直接抄录所提供 Survey 段落里的关键句，**不要改写**。每条 ≤120 字，必须保留原英文术语。
- synthesis 必须中文；除引用外不得编造新观点。
- 不要 markdown，不要其他多余文字。
"""

def call_qwen(system: str, user: str, max_retries: int = 3, timeout: int = 90) -> str:
    payload = json.dumps({
        "model": MODEL,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
    }).encode("utf-8")
    last_err = None
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(
                DASHSCOPE_URL,
                data=payload,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                body = json.loads(resp.read().decode("utf-8"))
                return body["choices"][0]["message"]["content"]
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
            last_err = e
            print(f"  [retry {attempt+1}/{max_retries}] {e}")
            time.sleep(2 + attempt * 3)
    raise RuntimeError(f"qwen call failed: {last_err}")

def build_user_prompt(b: dict, survey_passages: list[str]) -> str:
    lin = b.get("lineage") or {}
    rubrics_text = "\n".join(f"- {k}: {v}" for k, v in b["rubrics"].items())
    crit = lin.get("critiques_received") or []
    crit_text = "\n".join(f"- {c}" for c in crit) if crit else "（无）"
    self_lim = lin.get("self_acknowledged_limits") or "（无）"
    note = lin.get("note") or ""
    sp_text = "\n".join(f"- {p}" for p in survey_passages) if survey_passages else "（Survey 中未提及）"
    return f"""## benchmark
short_name: {b['short_name']}
cite: {b['cite']}
pillar/tag: {b.get('tag')}
year: {lin.get('year','')}
family: {lin.get('family','')}
note: {note}

## 已有 14-rubric 评论（仅作上下文）
{rubrics_text}

## lineage_map: critiques_received（原文）
{crit_text}

## lineage_map: self_acknowledged_limits（原文）
{self_lim}

## A_Comprehensive_Survey_of_AI_Ethics_Benchmarks 段落（原文，含上下文窗口）
{sp_text}

按上面 JSON schema 输出。"""

# --- 6. Main -----------------------------------------------------------------
def main():
    lineage = load_lineage()
    diag_rows = load_diagnosis()
    bench_rows = build_match_index(diag_rows, lineage)
    survey = load_survey_text()

    targets = [b for b in bench_rows if not b["is_survey"]]
    print(f"Loaded {len(bench_rows)} diag rows; targeting {len(targets)} benchmarks.")
    matched = [b for b in targets if b["lineage"]]
    print(f"  Matched to lineage: {len(matched)}; unmatched: {len(targets) - len(matched)}")

    results = {}
    if OUT_JSON.exists():
        results = json.loads(OUT_JSON.read_text())
        print(f"Resuming with {len(results)} cached")

    raw_fp = OUT_RAW.open("a", encoding="utf-8")
    for i, b in enumerate(targets, 1):
        key = b["short_name"]
        if key in results:
            print(f"[{i}/{len(targets)}] {key} cached")
            continue
        # Find Survey passages by short_name and other obvious aliases
        passages = find_survey_passages(survey, b["short_name"])
        if not passages and b.get("lineage"):
            passages = find_survey_passages(survey, b["lineage"].get("short_name", ""))
        prompt = build_user_prompt(b, passages)
        print(f"[{i}/{len(targets)}] {key} | lineage={'Y' if b['lineage'] else 'N'} | survey_passages={len(passages)}")
        try:
            raw = call_qwen(SYSTEM_PROMPT, prompt)
        except Exception as e:
            print(f"  FAILED: {e}")
            continue
        raw_fp.write(json.dumps({"short_name": key, "raw": raw}, ensure_ascii=False) + "\n")
        raw_fp.flush()
        try:
            parsed = json.loads(re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE))
        except Exception as e:
            print(f"  parse error: {e}")
            continue
        parsed["_meta"] = {
            "row": b["row"], "tag": b["tag"], "cite": b["cite"],
            "pillar": (b.get("lineage") or {}).get("_pillar"),
            "survey_passage_count": len(passages),
        }
        results[key] = parsed
        OUT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2))
    raw_fp.close()
    print(f"Done. {len(results)} entries → {OUT_JSON}")

if __name__ == "__main__":
    main()
