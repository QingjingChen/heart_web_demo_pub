"""Apply per_rubric.json to 科技伦理toolkit.xlsx:
  1. Benchmark诊断: APPEND lineage_relevant / external_relevant to each rubric cell
     (existing comments preserved; new content marked with leading separator).
  2. Tool来源分布表: REPLACE with new 14-rubric structure
     columns: 一级标签 | Benchmark/Source | 短名 | Source file/path |
              Representative tool | [14 rubric cells] |
              Adaptation example | What it improves | Lineage critiques (full) |
              Self limits | External views (full)
"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
TOOLKIT = ROOT.parent / "workbooks" / "科技伦理toolkit.xlsx"
PER_RUBRIC = ROOT / "per_rubric.json"
LINEAGE_VIEWS = ROOT / "lineage_views.json"

RUBRICS = [
    "Construct clarity",
    "Normative grounding",
    "Source provenance and evidence fitness",
    "Context and stakeholder coverage",
    "Real-world harm linkage",
    "Scenario validity",
    "Task-format fit",
    "Ground truth and disagreement design",
    "Metric validity",
    "Evaluator reliability",
    "Data and annotation QA",
    "Robustness against gaming and contamination",
    "Documentation and reproducibility",
    "Maintenance and update governance",
]
RUBRIC_ZH = {
    "Construct clarity": "构念清晰度",
    "Normative grounding": "规范基础",
    "Source provenance and evidence fitness": "来源可追溯",
    "Context and stakeholder coverage": "语境与主体覆盖",
    "Real-world harm linkage": "现实伤害连接",
    "Scenario validity": "场景真实性",
    "Task-format fit": "任务形式适配",
    "Ground truth and disagreement design": "答案与分歧设计",
    "Metric validity": "指标有效性",
    "Evaluator reliability": "评分器可靠性",
    "Data and annotation QA": "数据与标注质保",
    "Robustness against gaming and contamination": "防污染与防刷榜",
    "Documentation and reproducibility": "文档与可复现",
    "Maintenance and update governance": "时效性与动态演进",
}

LINEAGE_HEADER = "\n\n— lineage 引用 —\n"
EXTERNAL_HEADER = "\n\n— 外部观点 (Survey) —\n"


def format_cell_content(rubric_cell: dict, include_paper: bool = False) -> str:
    """Return string to append/insert into a rubric cell."""
    parts = []
    if include_paper and rubric_cell.get("paper_method"):
        parts.append(rubric_cell["paper_method"])
    lin = rubric_cell.get("lineage_relevant") or []
    if lin:
        parts.append(LINEAGE_HEADER.strip())
        for q in lin:
            parts.append(f"• {q}")
    ext = rubric_cell.get("external_relevant") or []
    if ext:
        parts.append(EXTERNAL_HEADER.strip())
        for q in ext:
            parts.append(f"• {q}")
    return "\n".join(parts)


def key_for_diag(cite: str) -> str:
    head = re.split(r"\s*[\(（]", str(cite))[0].strip()
    head = re.split(r"\s+/\s+", head)[0].strip()
    return head


def update_diagnosis(wb, per_rubric: dict, views: dict):
    """For each rubric cell in Benchmark诊断, append lineage/external from per_rubric.
    Preserves existing comment."""
    ws = wb["Benchmark诊断"]
    rubric_names = [ws.cell(row=2, column=c).value for c in range(2, 16)]

    # Build (bench_cite → per_rubric entry) for lookup
    per_by_cite = {v["_meta"]["bench_cite"]: v for v in per_rubric.values() if v.get("_meta")}
    per_by_short = {}
    for v in per_rubric.values():
        meta = v.get("_meta") or {}
        sn = meta.get("short_name_guess")
        if sn:
            per_by_short.setdefault(sn, []).append(v)

    updated = 0
    for r in range(3, ws.max_row + 1):
        cite = ws.cell(row=r, column=16).value
        if not cite:
            continue
        # Try exact cite match, then short_name
        entry = per_by_cite.get(str(cite).strip())
        if entry is None:
            head = key_for_diag(str(cite))
            cand = per_by_short.get(head, [])
            entry = cand[0] if cand else None
        if entry is None:
            continue
        # For each rubric column, append lineage/external
        for i, rn in enumerate(rubric_names):
            if rn not in entry:
                continue
            rcell = entry[rn]
            append = format_cell_content(rcell, include_paper=False)
            if not append:
                continue
            cur = ws.cell(row=r, column=2 + i).value
            new = (str(cur).rstrip() + "\n\n" + append) if cur else append
            ws.cell(row=r, column=2 + i).value = new
            ws.cell(row=r, column=2 + i).alignment = Alignment(wrap_text=True, vertical="top")
        updated += 1
    print(f"  Benchmark诊断: updated {updated} rows with lineage/external appendices")


def rebuild_tool_sheet(wb, per_rubric: dict, views: dict):
    """Drop the old Tool来源分布表 content and rewrite with 14-rubric layout.
    Preserves duplicates (same paper in different pillars: MoralChoice, TrustGPT)."""
    meta_rows_in_order = []
    if "Tool来源分布表" in wb.sheetnames:
        old = wb["Tool来源分布表"]
        for r in range(2, old.max_row + 1):
            cite = old.cell(row=r, column=2).value
            if not cite:
                continue
            meta_rows_in_order.append({
                "row": r,
                "tag": old.cell(row=r, column=1).value,
                "bench_cite": str(cite),
                "source_file": old.cell(row=r, column=3).value,
                "representative_tool": old.cell(row=r, column=4).value,
                "tool_explanation": old.cell(row=r, column=5).value,
                "what_improves": old.cell(row=r, column=6).value,
                "improved_rubrics_old": old.cell(row=r, column=7).value,
                "adaptation_example": old.cell(row=r, column=8).value,
            })
        wb.remove(old)

    ws = wb.create_sheet("Tool来源分布表", 0)

    # Build headers
    meta_headers = [
        "一级标签",
        "Benchmark / Source",
        "短名",
        "Source file/path",
        "Representative tool",
    ]
    rubric_headers = []
    for rn in RUBRICS:
        rubric_headers.append(f"{rn} ({RUBRIC_ZH[rn]})")
    trailing_headers = [
        "Adaptation example",
        "What it improves vs baseline (原文)",
        "Tool explanation (原文，全文)",
        "Improved rubrics (旧表标注)",
        "Lineage critiques (完整原文)",
        "Self acknowledged limits (原文)",
        "External views (Survey 原文)",
        "Synthesis (中文小结)",
    ]
    headers = meta_headers + rubric_headers + trailing_headers
    ws.append(headers)
    # Style header row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        if c <= len(meta_headers):
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        elif c <= len(meta_headers) + len(RUBRICS):
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
        else:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Build rows, ordered by original Tool sheet row position (preserves dups)
    for orig in meta_rows_in_order:
        cite = orig["bench_cite"]
        entry = per_rubric.get(cite)
        if entry is None:
            print(f"  skip R{orig['row']}: no per_rubric entry for {cite!r}")
            continue
        meta = entry.get("_meta", {})
        short_name = meta.get("short_name_guess") or ""
        # Lookup lineage_views for trailing columns
        v = views.get(short_name, {}) if views else {}
        row_vals = [
            orig.get("tag") or meta.get("tag"),
            cite,
            short_name,
            orig.get("source_file"),
            orig.get("representative_tool"),
        ]
        for rn in RUBRICS:
            rc = entry.get(rn, {})
            cell_text = format_cell_content(rc, include_paper=True)
            row_vals.append(cell_text)
        row_vals.extend([
            orig.get("adaptation_example"),
            orig.get("what_improves"),
            orig.get("tool_explanation"),
            orig.get("improved_rubrics_old"),
            "\n".join(f"• {q}" for q in (v.get("lineage_critiques") or [])),
            v.get("self_limits", ""),
            "\n".join(f"• {q}" for q in (v.get("external_views") or [])),
            v.get("synthesis", ""),
        ])
        ws.append(row_vals)

    # Column widths
    widths = [16, 30, 18, 30, 32] + [40] * len(RUBRICS) + [40, 40, 50, 30, 50, 40, 50, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "F2"
    print(f"  Tool来源分布表: rebuilt with {ws.max_row - 1} rows × {len(headers)} cols")


def main():
    per_rubric = json.loads(PER_RUBRIC.read_text(encoding="utf-8"))
    views = json.loads(LINEAGE_VIEWS.read_text(encoding="utf-8"))
    print(f"per_rubric entries: {len(per_rubric)}, views entries: {len(views)}")
    wb = openpyxl.load_workbook(TOOLKIT)
    update_diagnosis(wb, per_rubric, views)
    rebuild_tool_sheet(wb, per_rubric, views)
    wb.save(TOOLKIT)
    print(f"Saved → {TOOLKIT}")


if __name__ == "__main__":
    main()
