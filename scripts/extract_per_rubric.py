"""Per-rubric extraction: for each paper in Tool来源分布表 (and Benchmark诊断),
Qwen摘录每个 rubric 下的「本文方法原文」+ lineage 引用 + external view 引用。

输入:
  - 科技伦理toolkit.xlsx (Tool来源分布表 + Benchmark诊断 + 诊断rubrics说明)
  - scripts/lineage_views.json (已抽好的 lineage_critiques / self_limits / external_views)
  - scripts/lineage_map*.json (note / family / year)
输出:
  - scripts/per_rubric.json: { short_name: { rubric_name: {paper_method, lineage_relevant, external_relevant}}}
  - scripts/per_rubric.raw.jsonl

最高原则：所有字段必须是输入原文连续片段，不 paraphrase 不编造。
"""
from __future__ import annotations

import json
import re
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
TOOLKIT = ROOT.parent / "workbooks" / "科技伦理toolkit.xlsx"
LINEAGE_VIEWS = ROOT / "lineage_views.json"
OUT_JSON = ROOT / "per_rubric.json"
OUT_RAW = ROOT / "per_rubric.raw.jsonl"

KEY_FILE = Path.home() / ".dashscope_key"
API_KEY = KEY_FILE.read_text().strip()
URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
MODEL = "qwen-max"

RUBRICS = [
    "Construct clarity",
    "Normative grounding",
    "Source provenance and evidence fitness",
    "Context and stakeholder coverage",
    "Real-world harm linkage",
    "Scenario validity",
    "Task-format fit",
    "Ground truth and disagreement design",
    "Metric validity",
    "Evaluator reliability",
    "Data and annotation QA",
    "Robustness against gaming and contamination",
    "Documentation and reproducibility",
    "Maintenance and update governance",
]

RUBRIC_ZH = {
    "Construct clarity": "构念清晰度",
    "Normative grounding": "规范基础",
    "Source provenance and evidence fitness": "来源可追溯与证据适配",
    "Context and stakeholder coverage": "语境与主体覆盖",
    "Real-world harm linkage": "现实伤害连接",
    "Scenario validity": "场景真实性",
    "Task-format fit": "任务形式适配",
    "Ground truth and disagreement design": "答案与分歧设计",
    "Metric validity": "指标有效性",
    "Evaluator reliability": "评分器可靠性",
    "Data and annotation QA": "数据与标注质保",
    "Robustness against gaming and contamination": "防污染与防刷榜",
    "Documentation and reproducibility": "文档与可复现",
    "Maintenance and update governance": "时效性与动态演进",
}


def load_tool_rows():
    """Read Tool来源分布表 rows."""
    wb = openpyxl.load_workbook(TOOLKIT, data_only=True)
    ws = wb["Tool来源分布表"]
    rows = []
    for r in range(2, ws.max_row + 1):
        tag = ws.cell(row=r, column=1).value
        bench = ws.cell(row=r, column=2).value
        src = ws.cell(row=r, column=3).value
        rep = ws.cell(row=r, column=4).value
        explan = ws.cell(row=r, column=5).value
        improves = ws.cell(row=r, column=6).value
        improved_rubrics = ws.cell(row=r, column=7).value
        adaptation = ws.cell(row=r, column=8).value
        if not bench:
            continue
        head = re.split(r"\s*[\(（]", str(bench))[0].strip()
        head = re.split(r"\s+/\s+", head)[0].strip()
        rows.append({
            "row": r,
            "tag": tag,
            "bench_cite": str(bench),
            "short_name_guess": head,
            "source_file": src,
            "representative_tool": rep,
            "tool_explanation": explan,
            "what_improves": improves,
            "improved_rubrics": improved_rubrics,
            "adaptation_example": adaptation,
        })
    return rows


def load_diag_comments():
    """Read existing rubric comments from Benchmark诊断 keyed by short_name."""
    wb = openpyxl.load_workbook(TOOLKIT, data_only=True)
    ws = wb["Benchmark诊断"]
    rubric_names = [ws.cell(row=2, column=c).value for c in range(2, 16)]
    out = {}
    for r in range(3, ws.max_row + 1):
        cite = ws.cell(row=r, column=16).value
        if not cite:
            continue
        head = re.split(r"\s*[\(（]", str(cite))[0].strip()
        head = re.split(r"\s+/\s+", head)[0].strip()
        rubrics = {}
        for i, name in enumerate(rubric_names):
            v = ws.cell(row=r, column=2 + i).value
            if v:
                rubrics[name] = str(v)
        out.setdefault(head, []).append({"row": r, "rubrics": rubrics})
    return out


SYSTEM_PROMPT = """你是 AI 伦理 benchmark toolkit 的整理助手。

任务：为一个 benchmark/paper，将其「本文方法描述（Tool explanation / What it improves）」、「lineage_critiques（后续 benchmark 的批评原文）」、「external_views（Survey 段落原文）」按 14 个 rubric 维度拆分。

**最高原则：所有输出字段必须是输入文本中的连续片段或近似原句，绝不 paraphrase、绝不编造、绝不总结。**用户会逐字 check。

输入：
- bench_cite: benchmark 名
- tool_explanation: 本文方法描述（英文为主）
- what_improves: 本文相对 baseline 的改进
- lineage_critiques: 后续 benchmark 对它的批评（原句列表）
- self_limits: 作者自承局限（原句）
- external_views: A_Comprehensive_Survey 段落（原文）

输出严格 JSON：
{
  "Construct clarity": {
    "paper_method": "...tool_explanation/what_improves 中与本 rubric 相关的原句摘录；若无则 \\"\\"...",
    "lineage_relevant": ["...lineage_critiques 中与本 rubric 相关的原句..."],
    "external_relevant": ["...external_views 中与本 rubric 相关的原句..."]
  },
  "Normative grounding": {...},
  ...
  "Maintenance and update governance": {...}
}

规则：
1. 14 个 rubric key 必须齐全：Construct clarity / Normative grounding / Source provenance and evidence fitness / Context and stakeholder coverage / Real-world harm linkage / Scenario validity / Task-format fit / Ground truth and disagreement design / Metric validity / Evaluator reliability / Data and annotation QA / Robustness against gaming and contamination / Documentation and reproducibility / Maintenance and update governance.
2. paper_method 必须是 tool_explanation 或 what_improves 中**连续**的句子片段（可截断到关键句），保留原文术语；与该 rubric 无关则填 ""。
3. lineage_relevant 必须是 lineage_critiques 列表里的**完整原句**（含来源标注，如 "— Jin 2024 (KoBBQ)"），与该 rubric 不相关则不放。
4. external_relevant 必须是 external_views 中**连续**句子片段，与该 rubric 不相关则不放。
5. 一句话可以同时属于多个 rubric（如同一原文同时涉及 construct + normative grounding），允许重复。
6. 不要 markdown，不要其他多余文字。
"""


def call_qwen(system: str, user: str, max_retries: int = 3, timeout: int = 120) -> str:
    payload = json.dumps({
        "model": MODEL,
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
        "temperature": 0.05,
        "response_format": {"type": "json_object"},
    }).encode("utf-8")
    last_err = None
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(URL, data=payload, method="POST",
                headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                body = json.loads(resp.read().decode("utf-8"))
                return body["choices"][0]["message"]["content"]
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
            last_err = e
            print(f"  [retry {attempt+1}/{max_retries}] {e}")
            time.sleep(2 + attempt * 3)
    raise RuntimeError(f"qwen call failed: {last_err}")


def build_user_prompt(tool_row: dict, views_entry: dict | None) -> str:
    crit_list = (views_entry or {}).get("lineage_critiques") or []
    self_lim = (views_entry or {}).get("self_limits") or ""
    ext_list = (views_entry or {}).get("external_views") or []
    return f"""## benchmark
bench_cite: {tool_row['bench_cite']}
pillar: {tool_row.get('tag')}

## tool_explanation（本文方法描述，原文）
{tool_row.get('tool_explanation') or '（无）'}

## what_improves（相对baseline 的改进，原文）
{tool_row.get('what_improves') or '（无）'}

## improved_rubrics（用户已标注的主要相关 rubric，参考）
{tool_row.get('improved_rubrics') or '（无）'}

## adaptation_example（本文工具的适配示例）
{tool_row.get('adaptation_example') or '（无）'}

## lineage_critiques（原句列表）
{json.dumps(crit_list, ensure_ascii=False, indent=2) if crit_list else '[]'}

## self_limits（原句）
{self_lim}

## external_views（Survey 原文段落）
{json.dumps(ext_list, ensure_ascii=False, indent=2) if ext_list else '[]'}

按上述 schema 输出 JSON，14 个 rubric 齐全。"""


def main():
    tool_rows = load_tool_rows()
    views_all = json.loads(LINEAGE_VIEWS.read_text(encoding="utf-8"))
    # Build short_name → views entry
    views_by_short = {k: v for k, v in views_all.items()}

    print(f"Tool rows: {len(tool_rows)}")
    matched = sum(1 for t in tool_rows if t["short_name_guess"] in views_by_short)
    print(f"  Matched to lineage_views: {matched}; unmatched: {len(tool_rows) - matched}")

    results = {}
    if OUT_JSON.exists():
        results = json.loads(OUT_JSON.read_text(encoding="utf-8"))
        print(f"Resuming with {len(results)} cached")

    raw_fp = OUT_RAW.open("a", encoding="utf-8")
    for i, tool in enumerate(tool_rows, 1):
        key = tool["bench_cite"]
        if key in results:
            print(f"[{i}/{len(tool_rows)}] {key} cached")
            continue
        views_entry = views_by_short.get(tool["short_name_guess"])
        prompt = build_user_prompt(tool, views_entry)
        print(f"[{i}/{len(tool_rows)}] R{tool['row']} | {key[:50]} | lineage={'Y' if views_entry else 'N'}")
        try:
            raw = call_qwen(SYSTEM_PROMPT, prompt)
        except Exception as e:
            print(f"  FAILED: {e}")
            continue
        raw_fp.write(json.dumps({"key": key, "row": tool["row"], "raw": raw}, ensure_ascii=False) + "\n")
        raw_fp.flush()
        try:
            parsed = json.loads(re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE))
        except Exception as e:
            print(f"  parse error: {e}")
            continue
        # Ensure 14 rubrics exist
        for rn in RUBRICS:
            parsed.setdefault(rn, {"paper_method": "", "lineage_relevant": [], "external_relevant": []})
        parsed["_meta"] = {
            "row": tool["row"], "tag": tool["tag"], "bench_cite": tool["bench_cite"],
            "short_name_guess": tool["short_name_guess"], "has_lineage": views_entry is not None,
        }
        results[key] = parsed
        OUT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2))
    raw_fp.close()
    print(f"Done. {len(results)} entries → {OUT_JSON}")


if __name__ == "__main__":
    main()
