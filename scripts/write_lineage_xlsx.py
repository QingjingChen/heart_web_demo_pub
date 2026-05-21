"""Render scripts/lineage_views.json → 科技伦理toolkit_lineage与外部观点.xlsx.

One sheet: 一行一个 benchmark，列 = pillar / cite / 短名 / lineage_critiques /
self_limits / external_views / synthesis / 已有14-rubric评论摘要。
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent
TOOLKIT = ROOT.parent / "workbooks" / "科技伦理toolkit.xlsx"
VIEWS_JSON = ROOT / "lineage_views.json"
OUT_XLSX = ROOT.parent / "workbooks" / "科技伦理toolkit_lineage与外部观点.xlsx"

LINEAGE_FILES = [
    ROOT / "lineage_map.json",
    ROOT / "lineage_map_humancentered.json",
    ROOT / "lineage_map_privacy.json",
    ROOT / "lineage_map_safety.json",
    ROOT / "lineage_map_trustworthy.json",
]

def load_lineage_all():
    pillar_of_file = {
        "lineage_map.json": "fairness",
        "lineage_map_humancentered.json": "humancentered",
        "lineage_map_privacy.json": "privacy",
        "lineage_map_safety.json": "safety",
        "lineage_map_trustworthy.json": "trustworthy",
    }
    out = {}
    for f in LINEAGE_FILES:
        d = json.loads(f.read_text(encoding="utf-8"))
        pillar = pillar_of_file.get(f.name, "")
        for k, v in d.items():
            if k == "_doc" or not isinstance(v, dict) or v.get("_alias_of"):
                continue
            out[v.get("short_name", k)] = {**v, "_pillar": pillar}
    return out

def build_inverse_critic_map(all_lin):
    """For each B, list (A, critique_text) pairs where B critiqued A."""
    critic_to_targets = defaultdict(list)
    for bench, v in all_lin.items():
        for c in v.get("critiques_received", []):
            m = re.match(r"^([^:]+?):(.+)$", c)
            if not m:
                continue
            critic, critique_text = m.group(1).strip(), m.group(2).strip()
            critic_to_targets[critic].append((bench, critique_text))
    return critic_to_targets

PILLAR_LABEL = {
    "fairness": "Fairness & Inclusive",
    "humancentered": "Human-centered",
    "privacy": "Privacy",
    "safety": "Safety",
    "trustworthy": "Trustworthy",
}

def join_list(xs, sep="\n"):
    if not xs:
        return ""
    return sep.join(str(x) for x in xs)

def main():
    views = json.loads(VIEWS_JSON.read_text(encoding="utf-8"))
    wb_old = openpyxl.load_workbook(TOOLKIT, data_only=True)
    ws_diag = wb_old["Benchmark诊断"]
    rubric_names = [ws_diag.cell(row=2, column=c).value for c in range(2, 16)]

    def existing_rubrics(row_idx: int) -> str:
        out = []
        for i, name in enumerate(rubric_names):
            v = ws_diag.cell(row=row_idx, column=2 + i).value
            if v:
                out.append(f"[{name}] {str(v)[:250]}")
        return "\n".join(out)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lineage与外部观点"

    headers = [
        "一级标签",
        "Cite（原表）",
        "短名",
        "lineage_critiques（后续benchmark批评，原文）",
        "self_limits（作者自承局限，原文）",
        "external_views（Survey/Hancox-Li/Raji/Eriksson 等原文片段）",
        "synthesis（中文小结，引用为主）",
        "已有14-rubric评论摘要（仅作上下文）",
        "原表行号",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Order: by pillar then short_name
    def sort_key(item):
        short_name, data = item
        pillar = (data.get("_meta") or {}).get("pillar") or "zz"
        row = (data.get("_meta") or {}).get("row") or 9999
        return (pillar, row, short_name)

    for short_name, data in sorted(views.items(), key=sort_key):
        meta = data.get("_meta", {})
        pillar = meta.get("pillar")
        pillar_label = PILLAR_LABEL.get(pillar, pillar or "")
        row_idx = meta.get("row")
        ws.append([
            pillar_label,
            meta.get("cite", ""),
            data.get("短名", short_name),
            join_list(data.get("lineage_critiques") or []),
            data.get("self_limits", ""),
            join_list(data.get("external_views") or []),
            data.get("synthesis", ""),
            existing_rubrics(row_idx) if row_idx else "",
            row_idx,
        ])

    widths = [18, 32, 22, 70, 50, 70, 60, 70, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "C2"

    # --- Sheet 2: 互相引用图谱 -----------------------------------------------
    all_lin = load_lineage_all()
    critic_map = build_inverse_critic_map(all_lin)

    def match_critic_to_short(critic_label: str) -> str:
        """Try to map 'Smith 2022 (HolisticBias)' → 'HolisticBias'."""
        m = re.search(r"\(([^)]+)\)", critic_label)
        if m and m.group(1) in all_lin:
            return m.group(1)
        # Fallback: maybe the label itself is a short_name
        if critic_label in all_lin:
            return critic_label
        # Match by trailing year removed
        cand = re.sub(r"\s+\d{4}.*$", "", critic_label).strip()
        if cand in all_lin:
            return cand
        return ""

    ws2 = wb.create_sheet("互相引用图谱")
    h2 = [
        "Benchmark（被引方）",
        "Pillar",
        "年份",
        "被批评：lineage_critiques 中的后续 benchmark（原文+出处）",
        "曾批评：本 benchmark 在 critiques_received 中曾批评的更早 benchmark（原文）",
        "self_acknowledged_limits（原文）",
    ]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFE699")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    pillar_of = {}
    for short_name, v in all_lin.items():
        pillar_of[short_name] = v.get("_pillar", "")

    rows_for_sheet2 = []
    for short_name in sorted(all_lin.keys()):
        v = all_lin[short_name]
        # Received critiques – these are textual, attribute each
        received = v.get("critiques_received", []) or []
        received_text = "\n".join(f"- {c}" for c in received)
        # What this benchmark critiqued (from critic_map: who is "this benchmark" as critic)
        critiqued = []
        for critic_label, targets in critic_map.items():
            mapped = match_critic_to_short(critic_label)
            if mapped == short_name:
                for tgt_bench, tgt_text in targets:
                    critiqued.append(f"→ {tgt_bench}: {tgt_text}")
        critiqued_text = "\n".join(critiqued) if critiqued else ""
        self_lim = v.get("self_acknowledged_limits", "") or ""
        rows_for_sheet2.append([
            short_name,
            PILLAR_LABEL.get(v.get("_pillar", ""), v.get("_pillar", "")),
            v.get("year", ""),
            received_text,
            critiqued_text,
            self_lim,
        ])
    for row in rows_for_sheet2:
        ws2.append(row)

    widths2 = [22, 14, 8, 70, 70, 50]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for r in range(2, ws2.max_row + 1):
        for c in range(1, len(h2) + 1):
            ws2.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "B2"

    # --- Sheet 3: 说明 -------------------------------------------------------
    ws3 = wb.create_sheet("说明", 0)
    notes = [
        ["科技伦理toolkit_lineage与外部观点.xlsx"],
        [""],
        ["来源", "原始 科技伦理toolkit.xlsx 的 Benchmark诊断 sheet（121 行，14 rubric 评论）"],
        ["扩展", "scripts/lineage_map*.json（curated critiques_received + self_acknowledged_limits）"],
        ["", "A_Comprehensive_Survey_of_AI_Ethics_Benchmarks (6).pdf 段落原文"],
        ["", "诊断rubrics说明 Survey来源列：Raji/Eriksson/Jiao/Gallegos/Hancox-Li/Corrêa/Hagendorff/Xu/Zhong/Lyu&Du/Mbiazi/White/Reuel/Gu/BetterBench"],
        ["", ""],
        ["生成模型", "qwen-max（DashScope OpenAI-compatible endpoint）"],
        ["原则", "评论尽量保留外部文章原文：lineage_critiques 与 self_limits 来自 lineage_map 原句；"],
        ["", "external_views 直接抄录 Survey PDF 段落；synthesis 用引号引导外部观点。"],
        ["", "已有 14-rubric 评论保留不动，本表只做补充。"],
        ["", ""],
        ["Sheet 1: Lineage与外部观点", "一行一个 benchmark，9 列：标签/Cite/短名/lineage_critiques/self_limits/external_views/synthesis/原 14-rubric 摘要/原行号"],
        ["Sheet 2: 互相引用图谱", "对每个有 lineage 记录的 benchmark，列出：被哪些后续 benchmark 批评；本 benchmark 曾批评哪些更早 benchmark；作者自承局限"],
        ["", ""],
        ["统计", f"Lineage与外部观点 sheet 共 {ws.max_row - 1} 个 benchmark；互相引用图谱共 {ws2.max_row - 1} 个 benchmark"],
    ]
    for row in notes:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 100
    for r in range(1, ws3.max_row + 1):
        for c in range(1, 3):
            ws3.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    wb.save(OUT_XLSX)
    print(f"Saved → {OUT_XLSX}")
    print(f"  Sheet 'Lineage与外部观点': {ws.max_row - 1} rows")
    print(f"  Sheet '互相引用图谱':     {ws2.max_row - 1} rows")

if __name__ == "__main__":
    main()
